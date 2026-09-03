"""Planilla del sprint — Korvance, septiembre 2026.
Tres pestañas: Prospectos (la lista), Cálculo (la cuenta en vivo), Tablero (derivado, no se escribe).
Fórmulas de la hoja de números, secciones 2, 3 y 9.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

OUT = '/tmp/claude-0/-home-user-koe/2d10d8e8-69ce-5397-9887-ba74bb52f9e3/scratchpad/kit/planilla-sprint-korvance.xlsx'
ROWS = 200          # filas preparadas en Prospectos
EXAMPLE = 3         # fila de ejemplo (referencia de formato, no se cuenta)
FIRST = 4           # primera fila de datos reales

INK = '1A1713'
ACCENT = 'B85206'
HEAD_FILL = PatternFill('solid', fgColor='E9E4DB')
INPUT_FILL = PatternFill('solid', fgColor='FFF9E8')   # celdas para completar
CALC_FILL = PatternFill('solid', fgColor='F2EFE9')    # celdas calculadas
OK_FILL = PatternFill('solid', fgColor='E4F0E4')
BAD_FILL = PatternFill('solid', fgColor='FBE3D8')
DUE_FILL = PatternFill('solid', fgColor='FFF3E6')

thin = Side(style='thin', color='D8D0C3')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
BOTTOM = Border(bottom=Side(style='thin', color=ACCENT))

F = 'Arial'
def font(sz=10, b=False, color=INK, it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

wb = Workbook()

# ─────────────────────────────────────────────── PROSPECTOS
ws = wb.active
ws.title = 'Prospectos'

COLS = [
    ('Nombre', 20, 'in'), ('Negocio', 22, 'in'),
    ('Origen', 9, 'in'), ('Dolor', 26, 'in'),
    ('1er contacto', 12, 'in'), ('Último toque', 12, 'in'), ('Próximo toque', 13, 'in'),
    ('Estado', 17, 'in'),
    ('C', 7, 'in'), ('F', 7, 'in'), ('P', 7, 'in'), ('T', 9, 'in'),
    ('H', 7, 'in'), ('V', 8, 'in'),
    ('Total mensual', 13, 'calc'), ('Total × 6', 12, 'calc'), ('Nivel sugerido', 15, 'calc'),
    ('Facturación', 13, 'in'), ('Quién decide', 16, 'in'),
    ('Precio dicho', 12, 'in'), ('Medio de pago', 17, 'in'),
    ('Notas textuales', 46, 'in'),
]

ws['A1'] = 'PROSPECTOS — sprint de septiembre'
ws['A1'].font = font(13, True)
ws['D1'] = ('Amarillo: se completa a mano.  Gris: se calcula solo, no escribir.  '
            'La fila 3 es un ejemplo del formato y no se cuenta: cargá desde la fila 4.  '
            'La mañana arranca filtrando "Próximo toque" hasta hoy.')
ws['D1'].font = font(9, color='4A423A', it=True)

for i, (name, width, kind) in enumerate(COLS, start=1):
    c = ws.cell(row=2, column=i, value=name)
    c.font = font(9, True)
    c.fill = HEAD_FILL
    c.border = BOX
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[2].height = 30

# fila de ejemplo (formato esperado) + filas vacías
EJEMPLO = ['Ejemplo: Javier Ortega', 'Asesoría fiscal (Valencia)', 'F',
           '1 · Seguimiento de leads y respuesta lenta',
           date(2026, 9, 2), date(2026, 9, 2), date(2026, 9, 4), 'abierto',
           100, 0.20, 0.25, 200, 12, 9, None, None, None,
           '10–50k', 'Él con un socio', date(2026, 9, 3), 'Link tarjeta USD',
           'Perdió 2 clientes en agosto. El comercial se va en octubre.']

for r in range(EXAMPLE, FIRST + ROWS):
    vals = EJEMPLO if r == EXAMPLE else [None] * len(COLS)
    for i, (name, width, kind) in enumerate(COLS, start=1):
        c = ws.cell(row=r, column=i)
        c.font = font(10)
        c.border = BOX
        c.alignment = Alignment(vertical='top', wrap_text=(name == 'Notas textuales'))
        if kind == 'in':
            c.fill = INPUT_FILL
            c.value = vals[i - 1]
        else:
            c.fill = CALC_FILL
    # Total mensual = C*F*P*T + H*4,3*V   (hoja de números, sección 2)
    ws.cell(row=r, column=15).value = (
        f'=IF(COUNT(I{r}:L{r})+COUNT(M{r}:N{r})=0,"",'
        f'IFERROR(I{r}*J{r}*K{r}*L{r},0)+IFERROR(M{r}*4.3*N{r},0))')
    ws.cell(row=r, column=16).value = f'=IF(O{r}="","",O{r}*6)'
    # Nivel sugerido (hoja de números, sección 3)
    ws.cell(row=r, column=17).value = (
        f'=IF(P{r}="","",IF(R{r}="<10k","Todavía no",'
        f'IF(P{r}<1500,"Todavía no",IF(P{r}<3500,"Nivel 1","Nivel 2 o 3"))))')
    for col in (15, 16):
        ws.cell(row=r, column=col).number_format = '$#,##0;($#,##0);-'
    ws.cell(row=r, column=10).number_format = '0%'
    ws.cell(row=r, column=11).number_format = '0%'
    ws.cell(row=r, column=12).number_format = '$#,##0'
    ws.cell(row=r, column=14).number_format = '$#,##0'
    ws.cell(row=r, column=17).font = font(10, True)
    for col in (5, 6, 7, 20):
        ws.cell(row=r, column=col).number_format = 'yyyy-mm-dd'

for i in range(1, len(COLS) + 1):
    ws.cell(row=EXAMPLE, column=i).font = font(10, it=True, color='4A423A')
ws.cell(row=EXAMPLE, column=15).font = font(10, it=True, color='4A423A')

# listas desplegables
def dv(cells, options, prompt):
    d = DataValidation(type='list', formula1='"' + ','.join(options) + '"',
                       allow_blank=True, showDropDown=False)
    d.prompt = prompt
    d.promptTitle = 'Elegí una opción'
    ws.add_data_validation(d)
    d.add(cells)

rng = lambda col: f'{col}{EXAMPLE}:{col}{FIRST + ROWS - 1}'
dv(rng('C'), ['T', 'R', 'F', 'L'], 'T tibio · R referido · F frío · L landing')
dv(rng('D'), ['1 · Seguimiento de leads y respuesta lenta',
              '2 · Coordinación interna y traspasos',
              '3 · Atención al cliente y consultas repetidas',
              '4 · Carga de datos, reportes y facturación'], 'Los cuatro dolores del avatar')
dv(rng('H'), ['abierto', 'reservó', 'sesión hecha', 'propuesta enviada',
              'cerrado', 'estacionado', 'descalificado'], 'Estado del prospecto')
dv(rng('R'), ['<10k', '10–50k', '>50k'], 'Facturación mensual del negocio, en dólares')
dv(rng('U'), ['Link tarjeta USD', 'Transferencia LLC', 'Pesos MEP'], 'Medio de cobro')

# próximo toque vencido o de hoy → resaltado
ws.conditional_formatting.add(
    f'A{FIRST}:V{FIRST + ROWS - 1}',
    FormulaRule(formula=[f'AND($G{FIRST}<>"",$G{FIRST}<=TODAY(),$H{FIRST}<>"cerrado",$H{FIRST}<>"descalificado")'],
                fill=DUE_FILL, stopIfTrue=False))

ws.freeze_panes = 'C4'
ws.auto_filter.ref = f'A2:V{FIRST + ROWS - 1}'


# ─────────────────────────────────────────────── CÁLCULO
cs = wb.create_sheet('Cálculo')
cs.column_dimensions['A'].width = 34
for col, w in zip('BCDEFG', (13, 11, 11, 13, 13, 13)):
    cs.column_dimensions[col].width = w

def title(ws_, row, text, note=''):
    c = ws_.cell(row=row, column=1, value=text)
    c.font = font(11, True, ACCENT)
    c.border = BOTTOM
    for i in range(2, 8):
        ws_.cell(row=row, column=i).border = BOTTOM
    if note:
        n = ws_.cell(row=row, column=3, value=note)
        n.font = font(9, color='4A423A', it=True)

cs['A1'] = 'CÁLCULO EN VIVO — una sesión por vez'
cs['A1'].font = font(13, True)
cs['A2'] = ('Se completa mientras hablás. Amarillo: lo que dice el cliente. Gris: se calcula solo. '
            'Viene con el ejemplo cargado para que veas los controles en verde: borralo antes de la primera sesión real.')
cs['A2'].font = font(9, color='4A423A', it=True)

title(cs, 4, 'Las seis letras', 'hoja de números, sección 2')
LETTERS = [
    ('C — consultas por mes', 100, '0'),
    ('F — fracción que se pierde', 0.20, '0%'),
    ('P — conversión de las bien atendidas', 0.25, '0%'),
    ('T — ticket: primera venta o primer mes', 200, '$#,##0'),
    ('H — horas manuales por semana', 12, '0.0'),
    ('V — valor de esa hora (sueldo ÷ 170)', 9, '$#,##0.00'),
]
for i, (label, val, fmt) in enumerate(LETTERS):
    r = 5 + i
    cs.cell(row=r, column=1, value=label).font = font(10)
    c = cs.cell(row=r, column=2, value=val)
    c.font = font(11, True, '0000FF'); c.fill = INPUT_FILL; c.border = BOX; c.number_format = fmt

title(cs, 12, 'El número')
NUM = [
    ('Consultas perdidas por mes', '=B5*B6', '0.0'),
    ('Ventas en juego por mes', '=B12*B7', '0.0'),
    ('Fuga comercial mensual', '=B13*B8', '$#,##0'),
    ('Horas manuales mensuales', '=B9*4.3', '0.0'),
    ('Costo de las horas', '=B15*B10', '$#,##0'),
    ('TOTAL MENSUAL', '=B14+B16', '$#,##0'),
    ('TOTAL × 6 (lo que decide el nivel)', '=B17*6', '$#,##0'),
]
for i, (label, formula, fmt) in enumerate(NUM):
    r = 12 + i
    lab = cs.cell(row=r, column=1, value=label)
    val = cs.cell(row=r, column=2, value=formula)
    val.number_format = fmt; val.fill = CALC_FILL; val.border = BOX
    bold = label.startswith('TOTAL')
    lab.font = font(10, bold); val.font = font(12 if bold else 10, bold)

cs['A20'] = 'Nivel sugerido'
cs['A20'].font = font(10, True)
cs['B20'] = ('=IF(B18<1500,"Todavía no",IF(B18<3500,"Nivel 1 — desde US$1.500",'
             '"Nivel 2 — desde US$3.500 (Nivel 3 si además cambia procesos seguido, '
             'nadie va a ser dueño del sistema, o pide acompañamiento)"))')
cs['B20'].font = font(11, True, ACCENT); cs['B20'].fill = CALC_FILL; cs['B20'].border = BOX
cs['A21'] = 'Piso de cuota para Nivel 3 (2% de la facturación mensual)'
cs['A21'].font = font(9, color='4A423A')
cs['B21'] = '=IF(B22="","completá la facturación →",B22*0.02)'
cs['B21'].number_format = '$#,##0'; cs['B21'].fill = CALC_FILL; cs['B21'].border = BOX; cs['B21'].font = font(10)
cs['A22'] = 'Facturación mensual del cliente (para el piso del 2%)'
cs['A22'].font = font(9, color='4A423A')
cs['B22'] = 25000
cs['B22'].number_format = '$#,##0'; cs['B22'].fill = INPUT_FILL; cs['B22'].border = BOX; cs['B22'].font = font(10, color='0000FF')

title(cs, 24, 'El recorrido', 'ocho pasos: de la consulta al cobro')
for i in range(8):
    r = 25 + i
    cs.cell(row=r, column=1, value=f'Paso {i+1}').font = font(9, color='4A423A')
    c = cs.cell(row=r, column=2); c.fill = INPUT_FILL; c.border = BOX; c.font = font(10)
cs.merge_cells(start_row=25, start_column=2, end_row=25, end_column=7)
for i in range(1, 8):
    cs.merge_cells(start_row=25 + i, start_column=2, end_row=25 + i, end_column=7)
for i in range(8):
    for col in range(3, 8):
        cs.cell(row=25 + i, column=col).border = BOX
        cs.cell(row=25 + i, column=col).fill = INPUT_FILL

title(cs, 34, 'La tabla de fugas', 'el costo se calcula solo; las tres celdas de control tienen que dar 0')
TH = ['Dónde ocurre', 'Tipo', 'Veces/mes o hs/sem', 'Minutos c/u', 'Costo mensual', 'Qué lo resuelve', 'Prioridad']
for i, h in enumerate(TH, start=1):
    c = cs.cell(row=35, column=i, value=h)
    c.font = font(9, True); c.fill = HEAD_FILL; c.border = BOX
    c.alignment = Alignment(horizontal='center', wrap_text=True)
cs.row_dimensions[35].height = 28
EJ_FUGAS = [
    ('Consulta entra por WhatsApp y se contesta al otro día', 'comercial', 12, 'más de 24 h', 'Respuesta automática y aviso al que atiende', '1'),
    ('Presupuesto enviado sin segundo intento', 'comercial', 8, '—', 'Recordatorio a las 48 h con tarea asignada', '1'),
    ('Cargar consultas a mano en la planilla', 'horas', 4, '10 min c/u', 'Formulario que carga solo', '2'),
    ('Armar cada presupuesto desde cero', 'horas', 3, '15 min c/u', 'Plantilla con datos precargados', '2'),
    ('Contestar las mismas preguntas', 'horas', 3, '7 min c/u', 'Respuestas guardadas', '3'),
    ('Pasar datos a facturación', 'horas', 2, '—', 'Integración', 'no tocar'),
]
for i in range(6):
    r = 36 + i
    ej = EJ_FUGAS[i]
    for col in range(1, 8):
        c = cs.cell(row=r, column=col); c.border = BOX
        c.font = font(10, it=True, color='4A423A')
        c.fill = CALC_FILL if col == 5 else INPUT_FILL
    cs.cell(row=r, column=1).value = ej[0]
    cs.cell(row=r, column=2).value = ej[1]
    cs.cell(row=r, column=3).value = ej[2]
    cs.cell(row=r, column=4).value = ej[3]
    cs.cell(row=r, column=6).value = ej[4]
    cs.cell(row=r, column=7).value = ej[5]
    # comercial: veces/mes × P × T   ·   horas: hs/sem × 4,3 × V
    cs.cell(row=r, column=5).value = (
        f'=IF(C{r}="","",IF(B{r}="comercial",C{r}*$B$7*$B$8,IF(B{r}="horas",C{r}*4.3*$B$10,"")))')
    cs.cell(row=r, column=5).number_format = '$#,##0'
d = DataValidation(type='list', formula1='"comercial,horas"', allow_blank=True, showDropDown=False)
cs.add_data_validation(d); d.add('B36:B41')
d2 = DataValidation(type='list', formula1='"1,2,3,no tocar"', allow_blank=True, showDropDown=False)
cs.add_data_validation(d2); d2.add('G36:G41')

CTRL = [
    ('Control 1 — consultas repartidas menos C × F',
     '=IF(COUNT($C$36:$C$41)=0,"",ROUND(SUMIF($B$36:$B$41,"comercial",$C$36:$C$41)-B12,2))'),
    ('Control 2 — horas repartidas menos H',
     '=IF(COUNT($C$36:$C$41)=0,"",ROUND(SUMIF($B$36:$B$41,"horas",$C$36:$C$41)-B9,2))'),
    ('Control 3 — suma de la tabla menos el total',
     '=IF(COUNT($C$36:$C$41)=0,"",ROUND(SUM($E$36:$E$41)-B17,0))'),
]
for i, (label, formula) in enumerate(CTRL):
    r = 43 + i
    cs.cell(row=r, column=1, value=label).font = font(10)
    c = cs.cell(row=r, column=2, value=formula)
    c.font = font(11, True); c.border = BOX; c.number_format = '0.00;-0.00;0'
    cs.conditional_formatting.add(f'B{r}', FormulaRule(formula=[f'AND(B{r}<>"",B{r}=0)'], fill=OK_FILL))
    cs.conditional_formatting.add(f'B{r}', FormulaRule(formula=[f'AND(B{r}<>"",B{r}<>0)'], fill=BAD_FILL))
cs['C43'] = 'Verde = cuadra. Rojo = falta repartir algo; se pregunta en la sesión, no se completa después.'
cs['C43'].font = font(9, color='4A423A', it=True)

cs['A47'] = ('Prueba de que la planilla está bien armada: con el ejemplo cargado (100 · 20% · 25% · US$200 · 12 h · US$9) '
             'el TOTAL MENSUAL tiene que dar US$1.464 y el TOTAL × 6, unos US$8.784.')
cs['A47'].font = font(9, color='4A423A', it=True)

# ─────────────────────────────────────────────── TABLERO
ts = wb.create_sheet('Tablero')
ts.column_dimensions['A'].width = 42
for col in 'BCD':
    ts.column_dimensions[col].width = 14
ts['A1'] = 'TABLERO DEL SPRINT'
ts['A1'].font = font(13, True)
ts['A2'] = 'Todo se calcula desde Prospectos. Acá no se escribe nada.'
ts['A2'].font = font(9, color='4A423A', it=True)

title(ts, 4, 'Los mínimos del mes')
P = f"Prospectos!"
MET = [
    ('Contactos cargados', f'=COUNTA({P}$A${FIRST}:$A${FIRST+ROWS-1})', '', ''),
    ('Conversaciones de venta (precio dicho)', f'=COUNT({P}$T${FIRST}:$T${FIRST+ROWS-1})', 12, 'objetivo del mes'),
    ('Diagnósticos reservados', f'=COUNTIF({P}$H${FIRST}:$H${FIRST+ROWS-1},"reservó")', '', ''),
    ('Sesiones hechas', f'=COUNTIF({P}$H${FIRST}:$H${FIRST+ROWS-1},"sesión hecha")', '', ''),
    ('Propuestas enviadas', f'=COUNTIF({P}$H${FIRST}:$H${FIRST+ROWS-1},"propuesta enviada")', '', ''),
    ('Cerrados', f'=COUNTIF({P}$H${FIRST}:$H${FIRST+ROWS-1},"cerrado")', 2, 'objetivo del mes'),
]
for i, (label, formula, target, note) in enumerate(MET):
    r = 5 + i
    ts.cell(row=r, column=1, value=label).font = font(10)
    c = ts.cell(row=r, column=2, value=formula)
    c.font = font(12, True); c.fill = CALC_FILL; c.border = BOX; c.number_format = '0'
    if target:
        t = ts.cell(row=r, column=3, value=target); t.font = font(10, color='4A423A'); t.number_format = '0'
        ts.cell(row=r, column=4, value=note).font = font(9, color='4A423A', it=True)
        ts.conditional_formatting.add(f'B{r}', FormulaRule(formula=[f'B{r}>=C{r}'], fill=OK_FILL))

title(ts, 12, 'Hoy')
HOY = [
    ('Toques que vencen hoy o antes',
     f'=COUNTIFS({P}$G${FIRST}:$G${FIRST+ROWS-1},"<="&TODAY(),{P}$H${FIRST}:$H${FIRST+ROWS-1},"<>cerrado")'),
    ('Abiertos sin próximo toque cargado',
     f'=COUNTIFS({P}$A${FIRST}:$A${FIRST+ROWS-1},"<>",{P}$G${FIRST}:$G${FIRST+ROWS-1},"",{P}$H${FIRST}:$H${FIRST+ROWS-1},"abierto")'),
]
for i, (label, formula) in enumerate(HOY):
    r = 13 + i
    ts.cell(row=r, column=1, value=label).font = font(10)
    c = ts.cell(row=r, column=2, value=formula)
    c.font = font(12, True); c.fill = CALC_FILL; c.border = BOX; c.number_format = '0'

title(ts, 16, 'Por origen', 'la red tibia primero')
for i, (code, name) in enumerate([('T', 'Tibios'), ('R', 'Referidos'), ('F', 'Fríos'), ('L', 'Landing')]):
    r = 17 + i
    ts.cell(row=r, column=1, value=name).font = font(10)
    c = ts.cell(row=r, column=2, value=f'=COUNTIF({P}$C${FIRST}:$C${FIRST+ROWS-1},"{code}")')
    c.font = font(11); c.fill = CALC_FILL; c.border = BOX; c.number_format = '0'

title(ts, 22, 'Fuga total identificada')
ts['A23'] = 'Suma de "Total mensual" de los abiertos'
ts['A23'].font = font(10)
ts['B23'] = f'=SUMIFS({P}$O${FIRST}:$O${FIRST+ROWS-1},{P}$H${FIRST}:$H${FIRST+ROWS-1},"abierto")'
ts['B23'].font = font(12, True); ts['B23'].fill = CALC_FILL; ts['B23'].border = BOX
ts['B23'].number_format = '$#,##0'

ts['A26'] = ('La fila 3 de Prospectos es el ejemplo y no entra en ningún contador: '
             'los datos reales arrancan en la fila 4. Podés borrarla o dejarla.')
ts['A26'].font = font(9, color='4A423A', it=True)

for sheet in (ws, cs, ts):
    sheet.sheet_view.showGridLines = False

wb.save(OUT)
print('escrito', OUT)
